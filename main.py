# project for checking VIES VAT number from excell file and results writes to second column of the same file
# uses tkinter, openpyxl, zeep
# results are returned to queue and then written to file
import logging
import threading
import time
import tkinter as tk
from queue import Queue
from tkinter import filedialog
import openpyxl
from zeep import CachingClient, Transport


# show new window with readme.md inside
def readme():
    # create new window
    readme_window = tk.Toplevel()
    readme_window.title("Readme")
    readme_window.geometry("800x800")

    # put it in center of the screen
    readme_window.update_idletasks()
    width = readme_window.winfo_width()
    height = readme_window.winfo_height()
    x = (readme_window.winfo_screenwidth() // 2) - (width // 2)
    y = (readme_window.winfo_screenheight() // 2) - (height // 2)
    readme_window.geometry('{}x{}+{}+{}'.format(width, height, x, y))

    # create scrollbar
    scrollbar = tk.Scrollbar(readme_window)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    # create text widget
    text_widget = tk.Text(readme_window, yscrollcommand=scrollbar.set)
    text_widget.pack(expand=True, fill=tk.BOTH)
    # insert text

    # configure scrollbar
    scrollbar.config(command=text_widget.yview)
    # add close button
    close_button = tk.Button(readme_window, text="Close", command=readme_window.destroy)
    close_button.pack(side=tk.BOTTOM)
    # read file readme.md and show it with formating
    with open("readme.md", "r") as readme_file:

        for index, line in enumerate(readme_file):
            line.replace("**", "")
            header = 0
            if line.strip().startswith("###"):
                line = line.replace("###", "")
                header = 3
            if line.startswith("##"):
                line = line[2:]
                header = 2
            elif line.startswith("#"):
                line = line[1:]
                header = 1
            text_widget.insert(tk.END, line)  # insert line to text widget
            if header == 3:
                # Tag and style the text
                text_widget.tag_add("###", str(index + 1) + ".0", str(index + 1) + ".end")
            elif header == 2:
                # Tag and style the text
                text_widget.tag_add("##", str(index + 1) + ".0", str(index + 1) + ".end")
            elif header == 1:
                text_widget.tag_add("#", str(index + 1) + ".0", str(index + 1) + ".end")

        text_widget.tag_config("#", font=("Arial", 18, "bold"))
        text_widget.tag_config("##", font=("Arial", 14, "bold"))
        text_widget.tag_config("###", font=("Arial", 12, "bold"))


class ViesChecker:
    def __init__(self):

        self.queue = Queue()
        self.max_row = None
        self.sheet = None
        self.wb = None
        self.error_queue = Queue()
        self.running_queue = Queue()
        self.vat_list = []

        self.root = tk.Tk()
        self.root.title("VAT number checker")
        self.root.geometry("400x800")

        # put it in center of the screen
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry('{}x{}+{}+{}'.format(width, height, x, y))

        self.pick_file = tk.Button(self.root, text="Vyber soubor", command=self.__file_path_fce, height=2, width=10)
        self.pick_file.pack(padx=10, pady=10)
        self.file_path = tk.Label(self.root, text="Zde bude vypsána cesta k souboru.", bg="red", width=50)
        self.file_path.pack(padx=10, pady=10)

        self.start = tk.Button(self.root, text="Start", command=self.work_flow, height=2, width=10)
        self.start.pack(padx=10, pady=10)

        self.status_bar = tk.Label(self.root, text="Status: ", bg="white", width=50)
        self.status_bar.pack()
        self.status_bar_threads = tk.Label(self.root, text="Zpracováno: 0/0 vláken.", bg="white", width=50)
        self.status_bar_threads.pack()

        self.error_counter_label = tk.Label(self.root, text="Počet chyb: ", bg="white", width=50)
        self.error_counter_label.pack()

        height_of_list = 20
        list_frame = tk.Frame(self.root)
        list_frame.pack()
        left_frame = tk.Frame(list_frame)
        left_frame.pack(side=tk.LEFT)
        tk.Label(left_frame, text="Chyby").pack()
        self.error_list = tk.Listbox(left_frame, width=20, height=height_of_list)
        self.error_list.pack()

        # create button for readme
        readme_button = tk.Button(self.root, text="Návod", command=readme, height=2, width=10)
        readme_button.pack(padx=10, pady=10)

        # create close button
        close_button = tk.Button(self.root, text="Close", command=self.root.destroy, height=2, width=10)
        close_button.pack(padx=10, pady=10)

        right_frame = tk.Frame(list_frame)
        right_frame.pack(side=tk.RIGHT)
        tk.Label(right_frame, text="Ověřené DIČ.").pack()
        self.ok_list = tk.Listbox(right_frame, width=20, height=height_of_list)
        self.ok_list.pack()

        # if there is a config.txt file, it will be used as a default path
        try:
            with open("config.txt", "r") as f:
                self.file_path.config(text=f.read(), bg="green")
                self.__load_file()
        except FileNotFoundError:
            pass
        self.root.mainloop()

    def __load_file(self):
        # function for loading file
        try:
            self.wb = openpyxl.load_workbook(self.file_path.cget("text"))
            self.sheet = self.wb.active
            self.max_row = self.sheet.max_row
            self.vat_list = []
            for i in range(1, self.max_row + 1):
                # if second column is empty, it will be checked
                if self.sheet.cell(row=i, column=2).value is None:
                    cell = self.sheet.cell(row=i, column=1)
                    self.vat_list.append(cell.value)
                else:
                    if self.sheet.cell(row=i, column=2).value == "ERROR":
                        cell = self.sheet.cell(row=i, column=1)
                        self.vat_list.append(cell.value)
            # if there are no VAT numbers to check, it will be disabled
            if len(self.vat_list) == 0:
                self.start.config(state=tk.DISABLED)
                self.status_bar_threads.config(text="Soubor je již kompletně ověřen.")
            else:
                self.start.config(state=tk.NORMAL)
                self.status_bar_threads.config(text="Zpracováno: " + str(0) + "/" + str(len(self.vat_list)))
            self.status_bar.config(text="Status: Soubor načten.")
            self.status_bar.config(background="green")
        except Exception as e:
            self.status_bar.config(text="Status: Chyba při načítání souboru.")
            self.status_bar.config(background="red")
            logging.error(e)

    def __file_path_fce(self):
        path = filedialog.askopenfilename()
        self.file_path.config(text=path, bg="green")
        with open("config.txt", "w") as f:
            f.write(path)
            self.__load_file()

    def work_flow(self):
        """
        Main function for checking VAT numbers
        :return:
        """
        # disable buttons
        self.pick_file.config(state=tk.DISABLED)
        self.start.config(state=tk.DISABLED)
        # create list of threads
        threads = []
        for i in self.vat_list:
            status = [False, 10]
            thread = threading.Thread(target=check_vat,
                                      args=(i, self.queue, self.error_queue, self.running_queue, status))
            threads.append([thread, i, status])
            print("Thread " + i + " initialized")

        finished = 0
        # while there are still threads to
        while finished < len(threads):
            start_time = time.time()
            threads[finished][0].start()
            wait_time = 45
            # wait for thread to finish
            while threads[finished][2][0] is False and threads[finished][2][
                1] > 0 and time.time() - start_time < wait_time:
                left_count = threads[finished][2][1]
                # change color of status bar if there is less than 5 tries left or less than 20 seconds to timeout
                if left_count < 5 or int(wait_time - (time.time() - start_time)) < 20:
                    self.status_bar.config(background="orange")
                else:
                    self.status_bar.config(background="white")
                self.status_bar.config(text="Čekám na " + threads[finished][1] + "\n Počet zbývajících pokusů: " +
                                            str(left_count) + "\n Čas do timeoutu: " + str(
                    int(wait_time - (time.time() - start_time))) + "s")
                time.sleep(0.3)
                self.root.update()

            # add to listbox
            if threads[finished][2][0] is False:
                self.error_list.insert(0, threads[finished][1])
                self.error_counter_label.config(text="Počet chyb: " + str(self.error_list.size()))
            else:
                self.ok_list.insert(0, threads[finished][1])
            time.sleep(1)
            # if thread is still not finnished
            if threads[finished][2][0] is False:
                self.status_bar.config(text="Čekám na " + threads[finished][1] + "\n Zabijím výpočet. 10s do konce.")
                self.root.update()
                # joit thread with timeout
                threads[finished][0].join(timeout=10)
            finished += 1
            self.status_bar_threads.config(text="Zpracováno: " + str(finished) + "/" + str(len(threads)))
            self.root.update()

            self.save()
        self.status_bar.config(text="Status: Dokončeno.")
        self.pick_file.config(state=tk.NORMAL)

    def save(self):
        """
        Function for saving results to file
        While there is something in queue, it will be saved to file
        :return:
        """
        while not self.queue.empty():
            vat, result = self.queue.get()

            for i in range(1, self.max_row + 1):
                if self.sheet.cell(row=i, column=1).value == vat:
                    if result:
                        self.sheet.cell(row=i, column=2).value = "VALID"
                    else:
                        self.sheet.cell(row=i, column=2).value = "INVALID"
        while not self.error_queue.empty():
            vat = self.error_queue.get()
            for i in range(1, self.max_row + 1):
                if self.sheet.cell(row=i, column=1).value == vat:
                    self.sheet.cell(row=i, column=2).value = "ERROR"
        # save file to new file
        self.wb.save(self.file_path.cget("text"))


def check_vat(vat, queue, error_queue, running_queue, status_array=None):
    """
    Function for checking VAT number using SOAP API from EU website (https://ec.europa.eu/taxation_customs/vies/)
    Using SOAP API from zeep library (https://python-zeep.readthedocs.io/en/master/)
    :param vat: VAT number to check
    :param queue: Queue for valid VAT numbers
    :param error_queue: Queue for error VAT numbers
    :param running_queue: Queue for running threads
    :param status_array: Status array of thread
    :return:
    """
    if status_array is None:
        status_array = [False, 3]
    done = status_array[0]

    try:
        while status_array[1] > 0:
            try:
                running_queue.put(vat)
                # set timeout for connection
                transport = Transport(timeout=5)
                # create client for SOAP API using CachingClient
                client = CachingClient('http://ec.europa.eu/taxation_customs/vies/checkVatService.wsdl',
                                       transport=transport).service
                # get result from SOAP API
                result = client.checkVat(countryCode=vat[:2], vatNumber=vat[2:])
                # add to queue
                queue.put((vat, result["valid"]))
                # set status to done
                status_array[0] = True
                break
            except Exception as e:
                # if error, decrease number of tries
                status_array[1] -= 1
                logging.error(e)
                time.sleep(3)
                running_queue.get()
        if not status_array[0]:
            error_queue.put(vat)

    except Exception as e:
        error_queue.put(vat)
        print(e)
    finally:
        # remove from running queue
        running_queue.get()
        print("Thread " + vat + " finished")

        # print active threads


if __name__ == '__main__':
    ViesChecker()
